"""
Excel Transcript Parser
Extracts transcript data from Excel (.xlsx, .xls) files using openpyxl.
"""

import re
from typing import Optional, List, Dict
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

from models.transcript import Transcript, CourseRecord


class ExcelParser:
    """
    Parses Excel transcript files into Transcript objects.

    Handles:
    1. Structured spreadsheets with header rows
    2. Positional fallback (course code, credits, grade, semester columns)
    """

    COURSE_CODE_RE = re.compile(r"\b([A-Z]{2,4}\d{3,4}[A-Z]?)\b")
    GRADE_RE = re.compile(r"^(A|A-|B\+|B|B-|C\+|C|C-|D\+|D|F|W|I)$")
    SEMESTER_RE = re.compile(
        r"\b(Spring|Summer|Fall|Winter)\s+(\d{4})\b", re.IGNORECASE
    )

    HEADER_PATTERNS = {
        "course_code": re.compile(r"course[\s_]*(?:code|id|no)?", re.IGNORECASE),
        "credits": re.compile(r"credit(?:s)?|cr\.?|hours?", re.IGNORECASE),
        "grade": re.compile(r"grade|letter[\s_]*grade|mark", re.IGNORECASE),
        "semester": re.compile(r"semester|term|session|period", re.IGNORECASE),
    }

    @classmethod
    def parse_excel(
        cls,
        file_path: str,
        student_id: Optional[str] = None,
        student_name: Optional[str] = None,
        program: Optional[str] = None,
    ) -> Transcript:
        """
        Parse an Excel transcript file.

        Args:
            file_path: Path to .xlsx or .xls file
            student_id: Optional student ID
            student_name: Optional student name
            program: Optional program name

        Returns:
            Transcript object

        Raises:
            ImportError: If openpyxl is not installed
            FileNotFoundError: If file doesn't exist
            ValueError: If no valid records found
        """
        if load_workbook is None:
            raise ImportError(
                "openpyxl is required for Excel parsing.\n"
                "Install it with: pip install openpyxl"
            )

        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"Excel file not found: {file_path}")

        transcript = Transcript(
            student_id=student_id,
            student_name=student_name,
            program=program,
        )

        wb = load_workbook(file_path, read_only=True, data_only=True)
        records = []

        for sheet in wb.worksheets:
            sheet_records = cls._extract_from_sheet(sheet)
            records.extend(sheet_records)

        wb.close()

        if not records:
            raise ValueError(
                f"Could not extract transcript data from {file_path}.\n"
                f"Ensure the Excel file contains columns:\n"
                f"  Course Code, Credits, Grade, Semester"
            )

        for record in records:
            transcript.add_record(record)

        return transcript

    @classmethod
    def _extract_from_sheet(cls, sheet) -> List[CourseRecord]:
        """Extract records from a single worksheet."""
        records = []
        rows = list(sheet.iter_rows(values_only=True))

        if not rows or len(rows) < 2:
            return records

        # Try to identify columns from header row
        header_row = [str(cell or "").strip() for cell in rows[0]]
        col_map = cls._identify_columns(header_row)

        if col_map:
            # Structured extraction with identified headers
            for row in rows[1:]:
                cells = [str(cell or "").strip() if cell is not None else "" for cell in row]
                record = cls._parse_row_mapped(cells, col_map)
                if record:
                    records.append(record)
        else:
            # Positional fallback — try every row
            for row in rows:
                cells = [str(cell or "").strip() if cell is not None else "" for cell in row]
                record = cls._parse_row_positional(cells)
                if record:
                    records.append(record)

        return records

    @classmethod
    def _identify_columns(cls, header_row: list) -> Optional[Dict[str, int]]:
        """Try to identify column indices from a header row."""
        col_map = {}
        for idx, cell in enumerate(header_row):
            if not cell:
                continue
            for field, pattern in cls.HEADER_PATTERNS.items():
                if pattern.search(cell) and field not in col_map:
                    col_map[field] = idx
                    break

        if "course_code" in col_map and "grade" in col_map:
            return col_map
        return None

    @classmethod
    def _parse_row_mapped(
        cls, row: list, col_map: Dict[str, int]
    ) -> Optional[CourseRecord]:
        """Parse a row using identified column mapping."""
        try:
            course_code = row[col_map["course_code"]].strip()
            if not cls.COURSE_CODE_RE.match(course_code):
                return None

            grade = row[col_map.get("grade", -1)].strip()
            if not cls.GRADE_RE.match(grade):
                return None

            credits = 3.0
            if "credits" in col_map:
                try:
                    credits = float(row[col_map["credits"]] or 3)
                except (ValueError, TypeError):
                    credits = 3.0

            semester = "Unknown"
            if "semester" in col_map:
                semester = row[col_map["semester"]].strip() or "Unknown"

            return CourseRecord(
                course_code=course_code,
                credits=credits,
                grade=grade,
                semester=semester,
            )
        except (IndexError, ValueError, TypeError):
            return None

    @classmethod
    def _parse_row_positional(cls, row: list) -> Optional[CourseRecord]:
        """Parse a row without headers using positional guessing."""
        if len(row) < 3:
            return None

        course_code = None
        grade = None
        credits = 3.0
        semester = "Unknown"

        for cell in row:
            if not cell:
                continue
            if not course_code and cls.COURSE_CODE_RE.match(cell):
                course_code = cell
            elif not grade and cls.GRADE_RE.match(cell):
                grade = cell
            elif cls.SEMESTER_RE.search(cell):
                sem_m = cls.SEMESTER_RE.search(cell)
                semester = f"{sem_m.group(1).title()} {sem_m.group(2)}"
            else:
                try:
                    val = float(cell)
                    if 0 <= val <= 6:
                        credits = val
                except ValueError:
                    pass

        if course_code and grade:
            try:
                return CourseRecord(
                    course_code=course_code,
                    credits=credits,
                    grade=grade,
                    semester=semester,
                )
            except ValueError:
                return None
        return None
